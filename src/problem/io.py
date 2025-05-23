import os

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from pandas import DataFrame


def read_problem_from_excel(file_path: str,
                            cost_sheet: str = "Sheet1",
                            cost_range: str = None,
                            value_sheet: str = "Sheet1",
                            value_range: str = None,
                            ) -> dict:
    """
    최적화문제를 엑셀로부터 로드합니다.

    Args:
        file_path: 엑셀파일의 이름,
        cost_sheet: 비용 시트 이름
        cost_range: 비용 시트에서 읽을 데이터 범위, label을 포함합니다. ./data/data.xlsx을 참고하세요
        value_sheet: 가치 시트 이름
        value_range: 가치 시트에서 읽을 데이터 범위, label을 포함합니다. ./data/data.xlsx을 참고하세요
    Returns:
        Dict [DataFrame]: {"cost": 비용 데이터, "value": list[가치 데이터]}

    Raises:
        ValueError: 잘못된 엑셀 파일입니다.

    """
    if cost_range is None or value_range is None:
        raise ValueError("Please provide the range of cells to read.")

    wb = load_workbook(file_path, data_only=True)
    ws_cost = wb[cost_sheet]
    ws_value = wb[value_sheet]

    cost = read_cost_data(ws_cost, cost_range)
    value = read_value_data(ws_value, value_range)

    if cost.empty or not value:
        raise ValueError("잘못된 엑셀 파일입니다.")
    return {"cost": cost, "value": value}


def read_cost_data(ws: Workbook, value_range: str) -> pd.DataFrame:
    """
    엑셀 시트에서 비용 데이터를 읽어옵니다.
    Args:
        ws: 엑셀 시트
        value_range: 엑셀 시트에서 읽을 데이터 범위, label을 포함합니다. ex) "A1:J42"

    Returns:
        DataFrame: 비용 데이터
    """
    cost_range = ws[value_range.split(":")[0]: value_range.split(":")[1]]

    strategy_label = [cell.value for cell in cost_range[0]][1:]
    item_label = [record[0].value for record in cost_range][1:]

    costs = [[cell1.value, cell2.value, cell3.value] for _, cell1, cell2, cell3 in cost_range[1:]]
    costs = pd.DataFrame(costs, columns=strategy_label, index=item_label)

    return costs


def read_value_data(ws: Workbook, data_range: str) -> list[pd.DataFrame]:
    """
    엑셀 시트에서 가치 데이터를 읽어옵니다.
    Args:
        ws: 엑셀 시트
        data_range: 엑셀 시트에서 읽을 데이터 범위, label을 포함합니다. ex) "A1:J42"

    Returns:
        List[DataFrame]: 가치 데이터 리스트
    """
    value_range = ws[data_range.split(":")[0]: data_range.split(":")[1]]

    # 가치의 label을 분리하는 과정 ["고장률 민감도", "ENS 민감도", "CIC 민감도"] 와 같은 형태로 추출
    value_label = [cell.value for cell in value_range[0][1:]]  # 1행의 값들을 리스트로 저장
    value_label = [label for label in value_label if label is not None]  # None 값 제거
    value_dim = len(value_label)  # 가치의 차원

    # strategy 라벨을 분리하는 과정 ["교체", "정밀점검", "보통점검"] 와 같은 형태로 추출
    strategy_label = [cell.value for cell in value_range[1]]  # 2행의 값들을 리스트로 저장
    strategy_label = [label for label in strategy_label if label is not None]  # None 값 제거
    strategy_label = list(dict.fromkeys(strategy_label))  # 중복 제거
    num_strategy = len(strategy_label)  # 전략의 개수

    # item 라벨을 분리
    item_label = [record[0].value for record in value_range[2:]]
    num_item = len(item_label)  # 아이템의 개수

    values = []
    value_data = np.array([record[1:] for record in value_range[2:]])

    # 각 아이템에 대한 가치를 순회
    for i in range(num_item):
        item_info = []

        # 빈 행에 대해서는 건너뛰기
        if not value_data[i].any():
            continue

        # 각 가치에 대한 전략을 순회
        for j in range(value_dim):
            _value = []
            for k in range(num_strategy):
                _value.append(value_data[i][j * num_strategy + k].value)
            item_info.append(_value)
        values.append(pd.DataFrame(item_info, columns=strategy_label, index=value_label))

    return values


def write_solution_to_excel(file_path: str = "output_rel.xlsx",
                            sheet_name: str = "06. Maintenance Strategy",
                            start_cell: str = "A2",
                            problem: dict = None,
                            solution: list[int] | list[list[bool]] = None,
                            add_nothing: bool = True,
                            ) -> None:
    """
    최적화문제의 결과를 엑셀에 저장합니다. 시작 셀로부터 우하단으로 채워나갑니다. 행은 아이템을, 열은 전략을 나타냅니다.

    Args:
        file_path: 저장할 엑셀파일의 이름
        sheet_name: 엑셀 시트 이름
        start_cell: 시작 셀 위치, 해당 셀부터 우하단으로 채워나갑니다.
        problem: dict {"cost": DataFrame, "value": list[DataFrame]}
        solution: 각 아이템에 대해 선택된 전략 인덱스 또는 불리언 리스트
        add_nothing: True인 경우, '현상유지' 전략을 계산하여 추가합니다. 모든 전략을 선택하지 않은 경우, 비용과 가치가 0인 '현상유지' 전략으로 취급합니다.
    Returns:
        DataFrame: 비용 데이터
    """
    if problem is None or solution is None:
        raise ValueError("Please provide the problem and solution to write.")

    if solution[0] is list:
        solution = [solution[i].index(True) for i in range(len(solution))]

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name  # 첫 시트 이름 지정
        print(f"'{file_path}' 파일과 '{sheet_name}' 시트를 새로 생성했습니다.")
    else:
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            print(f"'{sheet_name}' 시트를 새로 생성했습니다.")

    ws = wb[sheet_name]


    strategy_label = problem["cost"].columns.tolist()
    item_label = problem["cost"].index.tolist()

    label_row = get_start_row(start_cell)
    start_row = label_row + 1
    label_col = get_start_col(start_cell)
    start_col = label_col + 1

    num_item = len(item_label)
    num_strategy = len(strategy_label)

    # 라벨 적용하기
    for row in range(0, num_item):
        ws.cell(row=row + label_row + 1, column=label_col, value=item_label[row])

    for col in range(0, num_strategy):
        ws.cell(row=label_row, column=col + label_col + 1, value=strategy_label[col])

    if not add_nothing:
        print("솔루션 저장에 '현상유지' 전략을 별도로 추가합니다.")
        ws.cell(row=label_row, column=label_col + num_strategy + 1, value="현상유지")
        for row in range(0, num_item):
            ws.cell(row=row + label_row + 1, column=label_col + num_strategy + 1, value=1 if solution[row] == -1 else 0)

    for i_item in range(0, num_item):
        for i_strategy in range(0, num_strategy):
            ws.cell(row=i_item + start_row, column=i_strategy + start_col,
                    value=1 if solution[i_item] == i_strategy else 0)

    wb.save(file_path)


def add_nothing_strategy(problem) -> dict:
    """
    문제 데이터에 아무것도 하지 않는 전략을 마지막에 추가합니다.
    Args:
        problem: dict {"cost": DataFrame, "value": list[DataFrame]}

    Returns:
        Dict: {"cost": 비용 데이터, "value": 가치 데이터 리스트}
    """
    # 마지막에 추가
    idx = len(problem["cost"].columns)  # 현재 컬럼 개수
    problem["cost"].insert(idx, "현상유지", 0)
    for i in range(len(problem["value"])):
        problem["value"][i].insert(idx, "현상유지", 0)

    return problem

def get_start_row(cell: str) -> int:
    """
    주어진 셀의 행 번호를 반환합니다.
    Args:
        cell (str): 셀 주소 (예: "A1")
    Returns:
        int: 행 번호
    """
    result = ""
    for char in cell:
        if char.isdigit():
            result += char

    return int(result)

def get_start_col(cell: str) -> int:
    """
    주어진 셀의 열 번호를 반환합니다.
    Args:
        cell (str): 셀 주소 (예: "A1")
    Returns:
        int: 열 번호
    """
    result = ""
    for char in cell:
        if char.isalpha():
            result += char

    return ord(result) - ord("A") + 1

def read_solution(file_path: str,
                  sheet_name: str = "06. Maintenance Strategy",
                    start_cell: str = "A2",
                    ) -> DataFrame:
    """
    엑셀 파일에서 솔루션을 읽어옵니다.
    :param file_path: 솔루션을 읽어올 엑셀 파일 경로
    :param sheet_name: 솔루션이 저장된 시트 이름
    :param start_cell: 솔루션이 시작되는 셀 주소, 해당 셀에서 우하단으로 값이 없는 셀까지 읽어옵니다.
    :return: DataFrame: 솔루션 데이터프레임
    """
    print(f"'{file_path}' 파일의 '{sheet_name}' 시트에서 솔루션을 읽어옵니다.")
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    start_row = get_start_row(start_cell)
    start_col = get_start_col(start_cell)
    data = []
    for row in ws.iter_rows(min_row=start_row, min_col=start_col):
        data.append([cell.value for cell in row])
    data = np.array(data)
    data = pd.DataFrame(data)
    data.columns = data.iloc[0]  # 첫 번째 행을 열 이름으로 설정
    data = data[1:]  # 첫 번째 행 제거
    data = data.reset_index(drop=True)  # 인덱스 초기화
    data.set_index(data.columns[0], inplace=True)  # 첫 번째 열을 인덱스로 설정
    return data