"""
이 모듈은 demo를 위해 임시적으로 작성하였기 때문에 별도의 주석을 제공하지 않습니다.
"""


import sys
import json

import openpyxl
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QLabel, QComboBox,
                               QLineEdit, QRadioButton, QPushButton, QTableWidget, QTableWidgetItem,
                               QVBoxLayout, QHBoxLayout, QGridLayout, QFrame, QMessageBox, QCheckBox)
from PySide6.QtCore import Qt
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

import main
from src.problem.io import read_solution

# 기본 설정값 (config.json이 없을 경우 사용)
DEFAULT_CONFIG = {
    "input": {
        "file_path": "data/200528_SK 계통(표준모델 적용).xlsm",
        "cost_range": "Z3:AC49",
        "cost_sheet": "04. reliability parameter for 3",
        "value_range": "A24:J71",
        "value_sheet": "05. results",
        "add_nothing_strategy": True
    },
    "solver": {
        "type": "SCIP",
        "problem_type": "cost_constraint",
        "cost_constraint": 1000,
        "value_weights": [1.0, 1.0, 1.0],
        "reliability_constraint": [150, 0.5, 0.5]
    },
    "output": {
        "file_path": "data/solution.xlsx",
        "sheet_name": "scip_cost_constraint",
        "cell": "A2"
    }
}


def load_config(config_path: str = 'configs/config.json'):
    """config.json 파일에서 설정을 로드합니다."""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"설정 파일 로드 오류: {e}, 기본 설정을 사용합니다.")
        return DEFAULT_CONFIG


def save_config(config: dict, config_path: str):
    """현재 설정을 config.json 파일에 저장합니다."""
    try:
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"설정 파일 저장 오류: {e}")
        raise e


class OptimizationUI(QMainWindow):
    def __init__(self):
        super().__init__()

        # config.json에서 설정 로드
        self.config_path = "configs/config.json"
        self.config = load_config()

        self.setWindowTitle("유지보수 전략 최적화")
        self.resize(900, 800)

        # UI 요소들
        # self.config_file_combo = QComboBox()
        # self.config_load_button = QPushButton("configs 새로고침")
        # self.solver_combo = QComboBox()
        # self.problem_type_combo = QComboBox()
        # self.cost_constraint_label = QLabel("비용 제약")
        # self.cost_constraint_input = QLineEdit()
        # self.sensitivity_weights_label = QLabel("민감도 가중치")
        # self.sensitivity_weights = []
        # self.sensitivity_constraint_frame = QFrame()
        # self.sensitivity_labels = ["고장율", "ENS", "CIC"]
        # self.sensitivity_constraints = []
        # self.add_nothing_checkbox = QCheckBox("문제 해결 시 '현상유지' 전략 추가")
        # self.sensitivity_radio = QRadioButton("민감도 미리보기")
        # self.cost_radio = QRadioButton("비용 미리보기")
        # self.sensitivity_frame = QFrame()
        # self.sensitivity_file_input = QLineEdit()
        # self.sensitivity_sheet_input = QLineEdit()
        # self.sensitivity_range_input = QLineEdit()
        # self.sensitivity_import_button = QPushButton("미리보기")
        # self.cost_frame = QFrame()
        # self.cost_file_input = QLineEdit()
        # self.cost_sheet_input = QLineEdit()
        # self.cost_range_input = QLineEdit()
        # self.cost_import_button = QPushButton("미리보기")
        # self.import_info = QLabel("")
        # self.output_file_input = QLineEdit()
        # self.output_sheet_input = QLineEdit()
        # self.output_cell_input = QLineEdit()
        # self.data_table = QTableWidget(8, 4)
        # self.save_button = QPushButton("설정 저장")
        # self.solve_button = QPushButton("풀이")
        # self.result_table = QTableWidget(9, 5)

        # 메인 위젯과 레이아웃 설정
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)

        # UI 컴포넌트 생성
        self.create_settings_section()
        self.create_output_section()
        self.create_import_section()
        self.create_data_table()
        self.create_buttons()
        self.create_result_section()

        # 문제와 솔루션 저장 변수
        self.problem = None
        self.solution = None

        # 설정 값을 UI에 적용
        self.apply_config_to_ui()

        # 상태 업데이트
        self.update_constraint_visibility()

    def on_config_file_selected(self, config_file):
        """드롭다운에서 설정 파일 선택 시 호출되는 메서드"""
        # 선택된 설정 파일 경로
        config_path = f"./configs/{config_file}"
        print(f"선택된 설정 파일: {config_path}")
        self.config_path = config_path

        # 설정 파일 로드
        self.config = load_config(config_path)

        # UI에 설정 적용
        self.apply_config_to_ui()

    def load_configs(self):
        # configs 폴더의 파일들 가져오기
        import os

        if os.path.exists("./configs"):
            self.config_file_combo.clear()
            config_files = [f for f in os.listdir("./configs") if f.endswith(".json")]
            self.config_file_combo.addItems(config_files)
        else:
            self.config_file_combo.addItem("config.json")

    def create_settings_section(self):
        # 문제 설정 섹션
        import_label = QLabel("설정 불러오기")
        import_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        self.main_layout.addWidget(import_label)

        configs_layout = QHBoxLayout()
        configs_layout.addWidget(QLabel("설정 파일:"))

        # 드롭다운 생성
        self.config_file_combo = QComboBox()

        self.load_configs()

        self.config_file_combo.showPopup()
        # 드롭다운 선택 이벤트 연결
        self.config_file_combo.currentTextChanged.connect(self.on_config_file_selected)

        configs_layout.addWidget(self.config_file_combo)

        self.config_load_button = QPushButton("configs 새로고침")
        self.config_load_button.setStyleSheet("background-color: #4CAF50; color: white; font-size: 14px;")
        self.config_load_button.clicked.connect(lambda: self.load_configs())
        configs_layout.addWidget(self.config_load_button)

        self.main_layout.addLayout(configs_layout)

        divider = QFrame()
        divider.setFrameShape(QFrame.HLine)
        divider.setFrameShadow(QFrame.Sunken)
        self.main_layout.addWidget(divider)

        settings_label = QLabel("문제 설정")
        settings_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        self.main_layout.addWidget(settings_label)

        settings_layout = QGridLayout()

        # 첫 번째 행 - 솔버, 문제 유형 선택
        settings_layout.addWidget(QLabel("문제해결 solver"), 0, 0)
        settings_layout.addWidget(QLabel("풀이할 문제"), 0, 1)
        settings_layout.addWidget(QLabel("제약 조건"), 0, 2, 1, 3)

        self.solver_combo = QComboBox()
        self.solver_combo.addItems(["SCIP", "CP-SAT"])
        settings_layout.addWidget(self.solver_combo, 1, 0)

        self.problem_type_combo = QComboBox()
        self.problem_type_combo.addItems(["비용 제약", "민감도 제약"])
        self.problem_type_combo.currentIndexChanged.connect(self.update_constraint_visibility)
        settings_layout.addWidget(self.problem_type_combo, 1, 1)

        # 비용 제약 필드
        self.cost_constraint_label = QLabel("비용 제약")
        settings_layout.addWidget(self.cost_constraint_label, 1, 2)
        self.cost_constraint_input = QLineEdit()
        self.cost_constraint_input.setText("1000")
        self.cost_constraint_input.setFixedWidth(150)  # 너비 고정
        settings_layout.addWidget(self.cost_constraint_input, 1, 3, 1, 3)

        # 민감도 가중치 필드
        self.sensitivity_weights_label = QLabel("민감도 가중치")
        self.sensitivity_weights_texts = ["고장율", "ENS", "CIC"]
        self.sensitivity_weights_labels = []
        settings_layout.addWidget(self.sensitivity_weights_label, 2, 2)
        self.sensitivity_weights_input = []
        for i, label in enumerate(self.sensitivity_weights_texts):
            weight_label = QLabel(label)
            settings_layout.addWidget(weight_label, 2, i + 3)
            weight_input = QLineEdit()
            weight_input.setText("1")
            weight_input.setFixedWidth(80)
            settings_layout.addWidget(weight_input, 3, i + 3)
            self.sensitivity_weights_input.append(weight_input)
            self.sensitivity_weights_labels.append(weight_label)

        self.normalization_label = QLabel("Min-Max 정규화:")
        settings_layout.addWidget(self.normalization_label, 4, 3)  # 행 번호는 기존 UI 구조에 맞게 조정
        self.normalization_checkbox = QCheckBox("모든 파라미터 정규화")
        self.normalization_checkbox.setToolTip(
            "모든 파라미터를 Min-Max정규화하여 민감도 목적함수를 구성합니다.")
        settings_layout.addWidget(self.normalization_checkbox, 4, 4, 1, 4)

        # 민감도 제약 프레임 (이름 변경)
        self.sensitivity_constraint_frame = QFrame()
        sensitivity_layout = QGridLayout(self.sensitivity_constraint_frame)
        sensitivity_layout.addWidget(QLabel("민감도 제약"), 0, 0)

        self.sensitivity_texts = ["고장율", "ENS", "CIC"]
        self.sensitivity_constraints = []

        for i, label in enumerate(self.sensitivity_texts):
            sensitivity_layout.addWidget(QLabel(label), 0, i + 1)
            constraint_input = QLineEdit()
            constraint_input.setText("1")
            constraint_input.setFixedWidth(80)
            sensitivity_layout.addWidget(constraint_input, 1, i + 1)
            self.sensitivity_constraints.append(constraint_input)

        settings_layout.addWidget(self.sensitivity_constraint_frame, 1, 2, 2, 4)

        settings_layout.addWidget(QLabel("현상유지 전략:"), 5, 0)  # 행 번호는 기존 UI 구조에 맞게 조정
        self.add_nothing_checkbox = QCheckBox("문제 해결 시 '현상유지' 전략 추가")
        self.add_nothing_checkbox.setToolTip(
            "체크하면 '현상유지' 전략이 추가되고, 체크하지 않으면 아무 전략도 선택하지 않을 경우 비용과 가치가 0인 '현상유지'로 취급합니다.")
        settings_layout.addWidget(self.add_nothing_checkbox, 5, 1, 1, 5)

        self.main_layout.addLayout(settings_layout)

    def create_import_section(self):
        # 값 불러오기 섹션
        divider = QFrame()
        divider.setFrameShape(QFrame.HLine)
        divider.setFrameShadow(QFrame.Sunken)
        self.main_layout.addWidget(divider)

        import_label = QLabel("값 미리보기")
        import_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        self.main_layout.addWidget(import_label)

        # 라디오 버튼
        radio_layout = QHBoxLayout()
        self.sensitivity_radio = QRadioButton("민감도 미리보기")
        self.sensitivity_radio.setChecked(True)
        self.cost_radio = QRadioButton("비용 미리보기")

        # 라디오 버튼 변경 이벤트 연결
        self.sensitivity_radio.toggled.connect(self.update_import_section)
        self.cost_radio.toggled.connect(self.update_import_section)

        radio_layout.addWidget(self.sensitivity_radio)
        radio_layout.addWidget(self.cost_radio)
        radio_layout.addStretch()
        self.main_layout.addLayout(radio_layout)

        # 민감도 입력 프레임
        self.sensitivity_frame = QFrame()
        sensitivity_layout = QHBoxLayout(self.sensitivity_frame)

        sensitivity_layout.addWidget(QLabel("파일명"))
        self.sensitivity_file_input = QLineEdit()
        self.sensitivity_file_input.setText(self.config["input"]["file_path"])
        sensitivity_layout.addWidget(self.sensitivity_file_input)

        sensitivity_layout.addWidget(QLabel("시트명"))
        self.sensitivity_sheet_input = QLineEdit()
        self.sensitivity_sheet_input.setText(self.config["input"]["value_sheet"])
        sensitivity_layout.addWidget(self.sensitivity_sheet_input)

        sensitivity_layout.addWidget(QLabel("범위 셀"))
        self.sensitivity_range_input = QLineEdit()
        self.sensitivity_range_input.setText(self.config["input"]["value_range"])
        sensitivity_layout.addWidget(self.sensitivity_range_input)

        self.sensitivity_import_button = QPushButton("미리보기")
        self.sensitivity_import_button.setStyleSheet("background-color: #4CAF50; color: white; font-size: 14px;")
        self.sensitivity_import_button.clicked.connect(lambda: self.load_data("sensitivity"))
        sensitivity_layout.addWidget(self.sensitivity_import_button)

        self.main_layout.addWidget(self.sensitivity_frame)

        # 비용 입력 프레임
        self.cost_frame = QFrame()
        cost_layout = QHBoxLayout(self.cost_frame)

        cost_layout.addWidget(QLabel("파일명"))
        self.cost_file_input = QLineEdit()
        self.cost_file_input.setText(self.config["input"]["file_path"])
        cost_layout.addWidget(self.cost_file_input)

        cost_layout.addWidget(QLabel("시트명"))
        self.cost_sheet_input = QLineEdit()
        self.cost_sheet_input.setText(self.config["input"]["cost_sheet"])
        cost_layout.addWidget(self.cost_sheet_input)

        cost_layout.addWidget(QLabel("범위 셀"))
        self.cost_range_input = QLineEdit()
        self.cost_range_input.setText(self.config["input"]["cost_range"])
        cost_layout.addWidget(self.cost_range_input)

        self.cost_import_button = QPushButton("미리보기")
        self.cost_import_button.setStyleSheet("background-color: #4CAF50; color: white; font-size: 14px;")
        self.cost_import_button.clicked.connect(lambda: self.load_data("cost"))
        cost_layout.addWidget(self.cost_import_button)

        self.main_layout.addWidget(self.cost_frame)

        self.import_info = QLabel("")
        self.import_info.setStyleSheet("font-size: 12px; color: gray;")
        self.main_layout.addWidget(self.import_info)

        # 초기 상태 설정
        self.update_import_section()

    def create_output_section(self):
        # 출력 설정 섹션
        divider = QFrame()
        divider.setFrameShape(QFrame.HLine)
        divider.setFrameShadow(QFrame.Sunken)
        self.main_layout.addWidget(divider)

        output_label = QLabel("출력 설정")
        output_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        self.main_layout.addWidget(output_label)

        output_layout = QGridLayout()

        # 출력 파일 경로
        output_layout.addWidget(QLabel("결과 파일 경로:"), 0, 0)
        self.output_file_input = QLineEdit()
        self.output_file_input.setText(self.config["output"]["file_path"])
        output_layout.addWidget(self.output_file_input, 0, 1, 1, 5)

        # 출력 시트 이름
        output_layout.addWidget(QLabel("결과 시트 이름:"), 1, 0)
        self.output_sheet_input = QLineEdit()
        self.output_sheet_input.setText(self.config["output"]["sheet_name"])
        output_layout.addWidget(self.output_sheet_input, 1, 1)

        output_layout.addWidget(QLabel("결과 입력 셀"), 1, 2)
        self.output_cell_input = QLineEdit()
        self.output_cell_input.setText(self.config["output"].get("cell", "A2"))
        output_layout.addWidget(self.output_cell_input, 1, 3)

        self.main_layout.addLayout(output_layout)

    def create_data_table(self):
        # 데이터 테이블
        self.data_table = QTableWidget(8, 4)
        self.data_table.setHorizontalHeaderLabels(["설비", "교체", "정밀점검", "단순점검"])

        # 예시 데이터 추가
        sample_data = [
            ["Sub_System1", "0.01", "10.26", "100,000,000"],
            ["Sub_System2", "0.01", "10.26", "100,000,000"],
            ["Sub_System3", "0.01", "10.26", "100,000,000"],
            ["Sub_System4", "0.01", "10.26", "100,000,000"],
            ["Sub_System5", "0.01", "10.26", "100,000,000"],
            ["Sub_System6", "0.01", "10.26", "100,000,000"],
            ["Sub_System7", "0.01", "10.26", "100,000,000"],
            ["Sub_System8", "0.01", "10.26", "100,000,000"]
        ]

        for i, row_data in enumerate(sample_data):
            for j, text in enumerate(row_data):
                self.data_table.setItem(i, j, QTableWidgetItem(text))

        self.data_table.resizeColumnsToContents()
        self.main_layout.addWidget(self.data_table)

    def create_buttons(self):
        # 버튼 레이아웃
        button_layout = QHBoxLayout()

        # 저장 버튼
        self.save_button = QPushButton("설정 저장")
        self.save_button.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px; font-size: 14px;")
        self.save_button.clicked.connect(lambda: self.save_current_config())
        button_layout.addWidget(self.save_button)

        # 풀이 버튼
        self.solve_button = QPushButton("풀이")
        self.solve_button.setStyleSheet("background-color: #0066cc; color: white; padding: 10px; font-size: 14px;")
        self.solve_button.clicked.connect(self.solve_problem)
        button_layout.addWidget(self.solve_button)

        self.main_layout.addLayout(button_layout)

    def create_result_section(self):
        # 결과 섹션
        result_label = QLabel("결과")
        result_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        self.main_layout.addWidget(result_label)

        self.result_table = QTableWidget(9, 5)
        self.result_table.setHorizontalHeaderLabels(["설비", "고장", "정밀점검", "보통점검", "현상유지"])

        # 예시 설비명 추가
        for i in range(9):
            self.result_table.setItem(i, 0, QTableWidgetItem(f"Sub_System{i + 1}"))

        self.result_table.resizeColumnsToContents()
        # 결과 테이블을 왼쪽에 놓고 결과 정보는 오른쪽에 배치
        result_layout = QHBoxLayout()

        # 왼쪽 결과 테이블
        result_layout.addWidget(self.result_table, 7)  # 비율 7

        # 오른쪽 결과 정보 프레임
        self.result_info_frame = QFrame()
        self.result_info_frame.setFrameShape(QFrame.StyledPanel)
        self.result_info_layout = QVBoxLayout(self.result_info_frame)

        # 결과 정보 레이블들
        total_cost_label = QLabel("비용: ")
        total_cost_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        self.total_cost = QLabel("0")
        total_value_label = QLabel("민감도: ")
        total_value_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        failure_label = QLabel("고장률")
        ens_label = QLabel("ENS")
        cic_label = QLabel("CIC")
        self.failure_value = QLabel("0")
        self.ens_value = QLabel("0")
        self.cic_value = QLabel("0")
        solution_time_label = QLabel("계산 시간: ")
        solution_time_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        self.elapsed_time = QLabel("0")

        # 정보 레이블 추가
        self.result_info_layout.addWidget(QLabel("최적화 결과 요약"), 0, Qt.AlignCenter)
        self.result_info_layout.addWidget(total_cost_label)
        self.result_info_layout.addWidget(self.total_cost)
        self.result_info_layout.addWidget(total_value_label)
        self.result_info_layout.addWidget(failure_label)
        self.result_info_layout.addWidget(self.failure_value)
        self.result_info_layout.addWidget(ens_label)
        self.result_info_layout.addWidget(self.ens_value)
        self.result_info_layout.addWidget(cic_label)
        self.result_info_layout.addWidget(self.cic_value)
        self.result_info_layout.addWidget(solution_time_label)
        self.result_info_layout.addWidget(self.elapsed_time)
        self.result_info_layout.addStretch()

        # 결과 레이아웃에 정보 프레임 추가
        result_layout.addWidget(self.result_info_frame, 3)  # 비율 3

        self.main_layout.addLayout(result_layout)


    def update_constraint_visibility(self):
        # 문제 유형에 따라 UI 요소 표시/숨김
        is_cost_constraint = self.problem_type_combo.currentText() == "비용 제약"

        # 비용 제약 관련 UI 요소
        self.cost_constraint_label.setVisible(is_cost_constraint)
        self.cost_constraint_input.setVisible(is_cost_constraint)

        # 민감도 가중치 관련 UI 요소 (비용 제약일 때만 표시)
        self.sensitivity_weights_label.setVisible(is_cost_constraint)
        for weight_input in self.sensitivity_weights_input:
            weight_input.setVisible(is_cost_constraint)

        for weight_label in self.sensitivity_weights_labels:
            weight_label.setVisible(is_cost_constraint)

        self.normalization_label.setVisible(is_cost_constraint)
        self.normalization_checkbox.setVisible(is_cost_constraint)

        # 민감도 제약 관련 UI 요소 (민감도 제약일 때만 표시)
        self.sensitivity_constraint_frame.setVisible(not is_cost_constraint)

        # 레이아웃 업데이트
        self.adjustSize()

    def update_import_section(self):
        """라디오 버튼 선택에 따라 입력 섹션 업데이트"""
        is_sensitivity = self.sensitivity_radio.isChecked()
        self.sensitivity_frame.setVisible(is_sensitivity)
        self.cost_frame.setVisible(not is_sensitivity)

    def apply_config_to_ui(self):
        """config 설정을 UI 컨트롤에 적용합니다."""
        # 솔버 설정
        solver_config = self.config.get("solver", {})
        solver_type = solver_config.get("type", "SCIP")
        problem_type = solver_config.get("problem_type", "cost_constraint")

        # 콤보박스 설정
        index = self.solver_combo.findText(solver_type)
        if index >= 0:
            self.solver_combo.setCurrentIndex(index)

        index = self.problem_type_combo.findText("비용 제약" if problem_type == "cost_constraint" else "민감도 제약")
        if index >= 0:
            self.problem_type_combo.setCurrentIndex(index)

        # 비용 제약 설정
        self.cost_constraint_input.setText(str(solver_config.get("cost_constraint", 1000)))
        self.normalization_checkbox.setChecked(solver_config.get("normalization", False))

        # 가중치 설정
        value_weights = solver_config.get("value_weights", [1.0, 1.0, 1.0])
        for i, weight in enumerate(value_weights):
            if i < len(self.sensitivity_weights_input):
                self.sensitivity_weights_input[i].setText(str(weight))

        # 민감도 제약 설정
        reliability_constraint = solver_config.get("reliability_constraint", [1, 1, 1])
        for i, constraint in enumerate(reliability_constraint):
            if i < len(self.sensitivity_constraints):
                self.sensitivity_constraints[i].setText(str(constraint))

        # 파일 경로 설정
        input_config = self.config.get("input", {})
        self.sensitivity_file_input.setText(input_config.get("file_path", ""))
        self.sensitivity_sheet_input.setText(input_config.get("value_sheet", ""))
        self.sensitivity_range_input.setText(input_config.get("value_range", ""))
        self.cost_file_input.setText(input_config.get("file_path", ""))
        self.cost_sheet_input.setText(input_config.get("cost_sheet", ""))
        self.cost_range_input.setText(input_config.get("cost_range", ""))

        input_config = self.config.get("input", {})
        self.add_nothing_checkbox.setChecked(input_config.get("add_nothing_strategy", True))

        output_config = self.config.get("output", {})
        self.output_file_input.setText(output_config.get("file_path", "data/solution.xlsx"))
        self.output_sheet_input.setText(output_config.get("sheet_name", "scip_cost_constraint"))
        self.output_cell_input.setText(output_config.get("cell", "A2"))

    def save_current_config(self):
        """현재 UI 설정을 config.json 파일에 저장합니다."""
        # 설정값 가져오기
        solver_type = self.solver_combo.currentText()
        problem_type = "cost_constraint" if self.problem_type_combo.currentText() == "비용 제약" else "reliability_constraint"

        # 설정 업데이트
        self.config["solver"]["type"] = solver_type
        self.config["solver"]["problem_type"] = problem_type

        # 문제 유형에 따라 다른 설정 저장
        if problem_type == "cost_constraint":
            self.config["solver"]["cost_constraint"] = float(self.cost_constraint_input.text())
            self.config["solver"]["value_weights"] = [float(w.text()) for w in self.sensitivity_weights_input]
            self.config["solver"]["value_normalization"] = self.normalization_checkbox.isChecked()
        else:
            self.config["solver"]["reliability_constraint"] = [float(c.text()) for c in self.sensitivity_constraints]

        # 입력 파일 경로 저장
        if self.sensitivity_radio.isChecked():
            self.config["input"]["file_path"] = self.sensitivity_file_input.text()
            self.config["input"]["value_sheet"] = self.sensitivity_sheet_input.text()
            self.config["input"]["value_range"] = self.sensitivity_range_input.text()
        else:
            self.config["input"]["file_path"] = self.cost_file_input.text()
            self.config["input"]["cost_sheet"] = self.cost_sheet_input.text()
            self.config["input"]["cost_range"] = self.cost_range_input.text()

        self.config["input"]["add_nothing_strategy"] = self.add_nothing_checkbox.isChecked()

        self.config["output"]["file_path"] = self.output_file_input.text()
        self.config["output"]["sheet_name"] = self.output_sheet_input.text()
        self.config["output"]["cell"] = self.output_cell_input.text()

        # 설정 저장
        success = save_config(self.config, self.config_path)
        if success:
            QMessageBox.information(self, "설정 저장", f"설정이 {self.config_path}에 저장되었습니다.")
        else:
            QMessageBox.warning(self, "저장 실패", "설정 저장에 실패했습니다.")

    def load_data(self, data_type):
        """데이터 불러오기 구현 - openpyxl 사용"""
        try:
            if data_type == "sensitivity":
                file_path = self.sensitivity_file_input.text()
                cell_range = self.sensitivity_range_input.text()
                sheet_name = self.sensitivity_sheet_input.text()
            else:  # cost
                file_path = self.cost_file_input.text()
                cell_range = self.cost_range_input.text()
                sheet_name = self.cost_sheet_input.text()

            self.import_info.setText(f"'{file_path}', '{sheet_name}' 시트의 {cell_range} 에서 {data_type}불러옵니다...")

            # 엑셀 파일 열기
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook[sheet_name]

            # 셀 범위 파싱 (예: "A1:D10")
            if cell_range:
                start_cell, end_cell = cell_range.split(":")
                start_col, start_row = coordinate_from_string(start_cell)
                end_col, end_row = coordinate_from_string(end_cell)

                start_col_idx = column_index_from_string(start_col)
                end_col_idx = column_index_from_string(end_col)
            else:
                # 범위가 지정되지 않은 경우 전체 데이터 영역 사용
                start_row, start_col_idx = 1, 1
                end_row = sheet.max_row
                end_col_idx = sheet.max_column

            # 헤더 추출 (첫 번째 행)
            headers = []
            for col in range(start_col_idx, end_col_idx + 1):
                cell_value = sheet.cell(start_row, col).value
                headers.append(str(cell_value) if cell_value is not None else f"Column {col}")

            # 테이블 설정 (헤더 제외한 데이터 행)
            row_count = end_row - start_row
            col_count = end_col_idx - start_col_idx + 1
            self.data_table.setRowCount(row_count)
            self.data_table.setColumnCount(col_count)
            self.data_table.setHorizontalHeaderLabels(headers)

            # 데이터 채우기 (헤더 행 제외)
            for row in range(start_row + 1, end_row + 1):
                table_row = row - start_row - 1
                for col in range(start_col_idx, end_col_idx + 1):
                    table_col = col - start_col_idx
                    value = sheet.cell(row, col).value
                    cell_value = str(value) if value is not None else ""
                    self.data_table.setItem(table_row, table_col, QTableWidgetItem(cell_value))

            self.data_table.resizeColumnsToContents()

            # 설정 업데이트
            if data_type == "sensitivity":
                self.config["input"]["file_path"] = file_path
                self.config["input"]["value_sheet"] = sheet_name
                self.config["input"]["value_range"] = cell_range
            else:
                self.config["input"]["file_path"] = file_path
                self.config["input"]["cost_sheet"] = sheet_name
                self.config["input"]["cost_range"] = cell_range

        except Exception as e:
            QMessageBox.critical(self, "오류", f"데이터 로드 오류: {e}")
            print(f"데이터 로드 오류: {e}")

    def solve_problem(self):
        # 현재 설정 저장
        self.save_current_config()

        try:
            solution, total_cost, total_value, solve_time = main.run_optimization(self.config_path)
            # 결과 표시
            self.total_cost.setText(f"{total_cost:.2f}")
            self.failure_value.setText(f"{total_value[0]:.10f}")
            self.ens_value.setText(f"{total_value[1]:.8f}")
            self.cic_value.setText(f"{total_value[2]:,.2f}")
            self.elapsed_time.setText(f"{solve_time:.2f}초")
            self.solution = solution

            self.display_solution()

            QMessageBox.information(self, "계산 완료",
                                    f"최적화 계산이 완료되었습니다.\n"
                                    f"총 비용: {total_cost}\n"
                                    f"계산 시간: {solve_time:.2f}초\n"
                                    f"결과가 {self.config['output']['file_path']} 파일에 저장되었습니다.")

        except Exception as e:
            QMessageBox.critical(self, "오류", f"최적화 문제 해결 오류: {e}")
            raise e

    def display_solution(self):
        # 결과 테이블 크기 설정
        self.result_table.setRowCount(len(self.solution))
        self.result_table.setColumnCount(5)  # 설비, 변경, 정밀점검, 보통점검, 현상유지
        self.result_table.setHorizontalHeaderLabels(["설비", "변경", "정밀점검", "보통점검", "현상유지"])

        solution = read_solution(file_path=self.config["output"]["file_path"],
                                 sheet_name=self.config["output"]["sheet_name"],
                                 start_cell=self.config["output"]["cell"])
        self.solution = solution

        # 솔루션 데이터프레임의 값을 결과 테이블에 표시
        for row_idx, row_name in enumerate(solution.index):
            # 설비 이름 설정
            self.result_table.setItem(row_idx, 0, QTableWidgetItem(str(row_name)))

            # 각 전략별 값 설정 (변경, 정밀점검, 보통점검, 현상유지)
            for col_idx, col_name in enumerate(solution.columns):
                self.result_table.setItem(row_idx, col_idx + 1, QTableWidgetItem(str(solution.iloc[row_idx, col_idx])))

        # 결과 테이블 크기 조정
        self.result_table.resizeColumnsToContents()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon("docs/icon.png"))
    window = OptimizationUI()
    window.show()
    sys.exit(app.exec())
